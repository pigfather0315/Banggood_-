import csv
import openpyxl


class BangGood(object):
    def __init__(self):
        self.original_file = '/Users/pigd/入口/codes/bg_data/审单文件/审单文件及结果10_29/Source/ES.csv'
        self.output_file = '/Users/pigd/入口/codes/bg_data/审单文件/审单文件及结果10_29/Result/ES_result_1029.xlsx'
        self.correct_file = '/Users/pigd/入口/codes/bg_data/审单文件/TT-UK-correct.xlsx'
        self.target_source = 'tradetracker'
        self.wait_status = ['back order', 'ready to process', 'payment confirmed', 'processing', 'payment preparing'
            , 'Contact Customer', 'ready to process', 'preorder', 'unverified order', 'system is processing', 'cod exception']
        self.cancel_status = ['cancel', 'order cancel', 'order cancelled', 'closed', 'payment pending',
                              'payment preparing', 'reject', 'expired', 'cod refused', 'grouping', ' grouped',
                              'built team', 'building team']
        self.done_status = ['cod received', 'shipped', 'order completed', 'accept', 'cod confirm']

    def write_data(self, sheet, line_num, data_list):
        for i, data in enumerate(data_list):
            sheet.cell(line_num + 1, i + 1, data)

    def construct_dict(self, path):
        data = {}
        workbook = openpyxl.load_workbook(path)
        worksheet = workbook.worksheets[0]
        rows = worksheet.max_row
        for one_row in worksheet.iter_rows(2, rows):
            if one_row[2].value is None:
                data[str(one_row[0].value)] = (one_row[1].value, '')
            else:
                data[str(one_row[0].value)] = (one_row[1].value, one_row[2].value)
        return data

    def evaluation(self):
        correct_data = self.construct_dict(self.correct_file)
        output_data = self.construct_dict(self.output_file)
        for key in correct_data.keys():
            corrects_status, correct_reason = correct_data[key][0], correct_data[key][1].lower()
            output_status, output_reason = output_data[key][0], output_data[key][1].lower()
            if corrects_status != output_status:
                print(key, output_status)
            else:
                if correct_reason != output_reason:
                    print(key, corrects_status, correct_reason)
                    print(key, output_status, output_reason)

    def processing(self):
        #### csv file ready
        f_r = open(self.original_file, 'r')
        reader = csv.reader(f_r)
        head_row = next(reader)
        ### xlsx file ready
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'order'
        self.write_data(worksheet, 0, ('Order', 'Status', 'Reason'))
        ### file writting
        for idx, row in enumerate(reader):
            order_number = row[0]
            source = row[1].strip()
            order_status = row[3].strip()
            if source != self.target_source:
                self.write_data(worksheet, idx + 1, [order_number, 'Cancel', 'Other Channel'])
            else:
                if order_status.lower() in self.wait_status:
                    self.write_data(worksheet, idx + 1, [order_number, 'Wait', ''])
                elif order_status.lower() in self.cancel_status:
                    self.write_data(worksheet, idx + 1, [order_number, 'Cancel', 'Order Cancelled'])
                elif order_status.lower() == 'refunded':
                    self.write_data(worksheet, idx + 1, [order_number, 'Cancel', 'Refunded'])
                elif order_status.lower() in self.done_status:
                    self.write_data(worksheet, idx + 1, [order_number, 'Done', ''])
                elif order_status.lower() == 'split':
                    self.write_data(worksheet, idx + 1, [order_number, 'Split', 'Manual operation needed'])

        workbook.save(filename=self.output_file)
        print('Processing completed')


if __name__ == '__main__':
    bg = BangGood()
    bg.processing()
    #bg.evaluation()
